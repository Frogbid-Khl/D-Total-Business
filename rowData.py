import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Folder containing Excel files
folder_path = 'canada/7.ontario/hamilton/'
output_file = folder_path+'combined_excel.xlsx'

# Columns to collect from each Excel file
desired_columns = ['Business URL', 'Business Name', 'Address', 'Website', 'Number', 'Business Open']

# Initialize an empty DataFrame to store all data
all_data = pd.DataFrame(columns=desired_columns)

# Initialize an empty set to store unique URLs
unique_urls = set()

# Iterate through each file in the folder
for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx'):
        # Read the Excel file
        file_path = os.path.join(folder_path, file_name)
        df = pd.read_excel(file_path)

        # Extract the data
        data = df[desired_columns]

        # Filter out rows with duplicate URLs
        data = data[~data['Business URL'].isin(unique_urls)]

        # Update the set of unique URLs
        unique_urls.update(data['Business URL'])

        # Exclude URLs starting with 'tel:'
        data = data[~data['Website'].str.startswith('tel:')]

        # Ensure Website column starts with 'http://' and replace 'https://www.bbb.org/' with 'empty'
        data['Website'] = data['Website'].apply(lambda x: 'empty' if x.startswith('https://www.bbb.org/') else ('http://' + x if not x.startswith('http') else x))

        # Concatenate the data with existing data
        all_data = pd.concat([all_data, data], ignore_index=True)

# Center align 'Business Open' column
all_data['Business Open'] = all_data['Business Open'].astype(str).str.center(20)


# Save the combined dataframe to an Excel file
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    all_data.to_excel(writer, index=False)

    # Access the workbook and worksheet objects
    workbook = writer.book
    worksheet = workbook.active

    # Set column widths
    column_widths = [500, 200, 300, 250, 150, 100]  # Widths in pixels
    for i, width in enumerate(column_widths, start=1):
        if width:
            column_letter = get_column_letter(i)
            worksheet.column_dimensions[column_letter].width = width / 7  # Convert pixels to character width

    # Set center alignment for the 5th column
    for row in worksheet.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Set wrap text for all cells
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrapText=True, horizontal='left')

print("Combined Excel file saved successfully.")
